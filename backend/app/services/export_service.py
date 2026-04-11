from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from app.models.project import Project
from app.models.task import Task, TaskLog

STATUS_LABELS = {
    "todo": "待办",
    "in_progress": "进行中",
    "blocked": "阻塞",
    "done": "已完成",
}

PRIORITY_LABELS = {
    "low": "低",
    "medium": "中",
    "high": "高",
    "urgent": "紧急",
}

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1E293B")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
BORDER_SIDE = Side(style="thin", color="E2E8F0")
CELL_BORDER = Border(
    left=BORDER_SIDE, right=BORDER_SIDE,
    top=BORDER_SIDE, bottom=BORDER_SIDE,
)


def _style_header_row(ws, row: int, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = CELL_BORDER


def _style_data_row(ws, row: int, col_count: int, alt: bool = False):
    bg = "F8FAFC" if alt else "FFFFFF"
    fill = PatternFill(fill_type="solid", fgColor=bg)
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.border = CELL_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def _set_col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def generate_excel(db: Session, project: Project) -> bytes:
    wb = Workbook()

    tasks = db.query(Task).filter(Task.project_id == project.id).all()
    total = len(tasks)
    done = sum(1 for t in tasks if t.status.value == "done")
    blocked = sum(1 for t in tasks if t.status.value == "blocked")
    in_progress = sum(1 for t in tasks if t.status.value == "in_progress")

    # ── Sheet 1: 项目概览 ──────────────────────────────────
    ws1 = wb.active
    ws1.title = "项目概览"
    ws1.row_dimensions[1].height = 22

    overview_headers = ["字段", "内容"]
    ws1.append(overview_headers)
    _style_header_row(ws1, 1, 2)

    overview_rows = [
        ("项目名称", project.name),
        ("项目状态", {"active": "进行中", "completed": "已完成", "archived": "已归档"}.get(project.status.value, project.status.value)),
        ("创建时间", str(project.created_at)[:16]),
        ("总任务数", total),
        ("待办", total - done - blocked - in_progress),
        ("进行中", in_progress),
        ("阻塞", blocked),
        ("已完成", done),
        ("完成率", f"{round(done / total * 100)}%" if total else "0%"),
    ]
    for i, row in enumerate(overview_rows, 2):
        ws1.append(list(row))
        _style_data_row(ws1, i, 2, alt=(i % 2 == 0))
    _set_col_widths(ws1, [18, 30])

    # ── Sheet 2: 任务明细 ──────────────────────────────────
    ws2 = wb.create_sheet("任务明细")
    ws2.row_dimensions[1].height = 22
    headers2 = ["任务标题", "描述", "负责人", "状态", "优先级", "进度(%)", "截止日期", "创建时间", "最后更新"]
    ws2.append(headers2)
    _style_header_row(ws2, 1, len(headers2))

    for i, t in enumerate(tasks, 2):
        ws2.append([
            t.title,
            t.description or "",
            t.assignee.name if t.assignee else "未指派",
            STATUS_LABELS.get(t.status.value, t.status.value),
            PRIORITY_LABELS.get(t.priority.value, t.priority.value),
            t.progress,
            str(t.due_date) if t.due_date else "",
            str(t.created_at)[:16],
            str(t.updated_at)[:16],
        ])
        _style_data_row(ws2, i, len(headers2), alt=(i % 2 == 0))
    _set_col_widths(ws2, [28, 36, 12, 10, 8, 8, 12, 16, 16])

    # ── Sheet 3: 进展日志 ──────────────────────────────────
    ws3 = wb.create_sheet("进展日志")
    ws3.row_dimensions[1].height = 22
    headers3 = ["任务标题", "记录人", "状态", "进度(%)", "进展内容", "记录时间"]
    ws3.append(headers3)
    _style_header_row(ws3, 1, len(headers3))

    row_idx = 2
    for t in tasks:
        logs = (
            db.query(TaskLog)
            .filter(TaskLog.task_id == t.id)
            .order_by(TaskLog.created_at)
            .all()
        )
        for log in logs:
            ws3.append([
                t.title,
                log.user.name,
                STATUS_LABELS.get(str(log.status), str(log.status)),
                log.progress,
                log.content,
                str(log.created_at)[:16],
            ])
            _style_data_row(ws3, row_idx, len(headers3), alt=(row_idx % 2 == 0))
            row_idx += 1

    if row_idx == 2:
        ws3.append(["暂无进展记录", "", "", "", "", ""])
        _style_data_row(ws3, 2, len(headers3))

    _set_col_widths(ws3, [28, 12, 10, 8, 48, 16])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
